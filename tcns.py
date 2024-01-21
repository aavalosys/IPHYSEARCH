import mysql.connector
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScaleRule
#from openpyxl.formatting.rule import ColorScale, FormatObject
import datetime
import pytz


def convertTuple(tup):
    st = ''.join(map(str, tup))
    return st

def tcnreport():
    mydb = mysql.connector.connect(
      host="10.10.26.4",
      user="readonly",
      password="",
      #database="fy2023w09"
      database="TCN"
    )
    mycursor = mydb.cursor()
    arreglo = []
    pelasthour = []
    guatemala_timezone = pytz.timezone('America/Guatemala')
    current_datetime = datetime.datetime.now(guatemala_timezone)
    year = current_datetime.strftime("%Y")
    month = current_datetime.strftime("%m")
    day = current_datetime.strftime("%d")
    hour = current_datetime.strftime("%H")
    nombrearchivo = hour + " Horas " + day + "-" + month + "-" + year
    #print (nombrearchivo)
    instanciaip = ""
    noop = 0
    arregloipinstancia = []
    query = "select * from stptcn order by hourrecolected desc"
    mycursor.execute(query)
    myresult = mycursor.fetchall()
    index = 0
    temportal = []
    for x in myresult:
        instanciaip = x[0] + "," + x[1]
        if instanciaip in arregloipinstancia:
            #print("x")
            #print (x)
            #print ("instancias")
            #print (instanciaip)
            #print ("arreglo")
            #print (arregloipinstancia)
            ############si esta#######
            index = arregloipinstancia.index(instanciaip)
            #print (index)
            #print ("que tiene 7")
            #print (arreglo[7])
            temportal = arreglo[index]
            if len(temportal) > 11:
                noop = 0
            else:
                temportal.append(x)
                arreglo[index] = temportal
        else:
            ############ no esta
            arregloipinstancia.append(instanciaip)
            temportal = []
            temportal.append(x)
            arreglo.append(temportal)
        #print (x)
   # print (arreglo)
   # c = 0
   # while c != len(arreglo):
   #     print (len(arreglo[c]))
   #     c = c + 1
   # print (len(arreglo))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resumen"
    ##################### ya se tiene data ordenada
    arregloanalisis = []
    c = 0
    temp = []
    recentvalue = ""
    oldestvalue = ""
    recentvaluen = 0
    oldestvaluen = 0
    valuefinal = 0
    contadorfila = 2
    contadorcolumna = 1
    cont = 0
    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 25
    sheet.column_dimensions["C"].width = 15
    sheet.column_dimensions["D"].width = 15
    sheet.column_dimensions["E"].width = 25
    sheet.column_dimensions["F"].width = 20
    sheet.column_dimensions["G"].width = 25
    sheet.column_dimensions["H"].width = 25
    sheet[f'{"A"}{1}'] = "Delta TCNs Ultimas 12hrs"
    sheet[f'{"B"}{1}'] = "Delta TCNs Ultima hora"
    sheet[f'{"C"}{1}'] = "IP PE"
    sheet[f'{"D"}{1}'] = "Instancia"
    sheet[f'{"E"}{1}'] = "Ultimo TCN"
    sheet[f'{"F"}{1}'] = "Interfaz"
    sheet[f'{"G"}{1}'] = "Ultima recolecion GMT"
    sheet[f'{"H"}{1}'] = "Ingeniero Asignado"
    sheet.auto_filter.ref = "A1:H1"
    clarofill = PatternFill(fill_type="solid", fgColor="FF0000")  # Green color fill
    sheet["A1"].fill = clarofill
    sheet["B1"].fill = clarofill
    sheet["C1"].fill = clarofill
    sheet["D1"].fill = clarofill
    sheet["E1"].fill = clarofill
    sheet["F1"].fill = clarofill
    sheet["G1"].fill = clarofill
    sheet["H1"].fill = clarofill
    bold_font = Font(bold=True)
    white_font = Font(color="FFFFFF")
    sheet["A1"].font = bold_font
    sheet["B1"].font = bold_font
    sheet["C1"].font = bold_font
    sheet["D1"].font = bold_font
    sheet["E1"].font = bold_font
    sheet["G1"].font = bold_font
    sheet["H1"].font = bold_font
    sheet["F1"].font = bold_font
    sheet["A1"].font = white_font
    sheet["B1"].font = white_font
    sheet["C1"].font = white_font
    sheet["D1"].font = white_font
    sheet["E1"].font = white_font
    sheet["F1"].font = white_font
    sheet["G1"].font = white_font
    sheet["H1"].font = white_font
   # FF0000
    color_scale_rule = ColorScaleRule(start_type='min', start_color='00FF00', end_type='max', end_color='FF0000')
    #differential_style = DifferentialStyle(fill=PatternFill(start_color="FF0000", mac="FF0000"))  # Relleno rojo
    #rule = Rule(type="expression", dxf=differential_style)
    #rule.formula = ['A2>20']
    #columna = "A" + contadorfila
    sheet.conditional_formatting.add('A2:A360', color_scale_rule)
    sheet.conditional_formatting.add('B2:B360', color_scale_rule)    
    #for row_index, row_data in enumerate(arreglo[0], 1):
    #    for col_index, cell_data in enumerate(row_data, 1):
    #        col_letter = get_column_letter(col_index)
    #        sheet[f'{col_letter}{row_index}'] = cell_data
    #workbook.save('datatcn.xlsx')
    while c != len(arreglo):
        temportal = arreglo[c]
        #### get first val
        if len(temportal) > 1:
            temp = temportal[0]
            recentvalue = temp[2]
            temp = temportal[len(temportal) - 1]
            oldestvalue = temp[2]
            temp = temportal[1]
            secondvalue = temp[2]
            #print (recentvalue + "-" + oldestvalue)
            oldestvaluen = int(oldestvalue)
            recentvaluen = int(recentvalue)
            secondvaluen = int(secondvalue)
            valuefinal = recentvaluen - oldestvaluen
            valuefinals = recentvaluen - secondvaluen
            stringtemp = convertTuple(temp[0]) + "," + convertTuple(temp[3])
            #print (stringtemp)
            #print (valuefinal)
            if valuefinal > 0  and not stringtemp in pelasthour:
                pelasthour.append(stringtemp)
                contadorcolumna = 1
                aux = 0
                col_letter = get_column_letter(contadorcolumna)
                sheet[f'{col_letter}{contadorfila}'] = valuefinal
                col_letter = get_column_letter(2)
                sheet[f'{col_letter}{contadorfila}'] = valuefinals
                contadorcolumna = contadorcolumna + 2
                implicitcount = 0
                arregloprovisional = arreglo[c]
                arregloprovisional = arregloprovisional[0]
                while implicitcount != len(arregloprovisional):
                    col_letter = get_column_letter(contadorcolumna - aux)
                    if implicitcount == 2:
                        col_letter = get_column_letter(contadorcolumna - 1)
                        aux = 1
                    else:
                        sheet[f'{col_letter}{contadorfila}'] = convertTuple(arregloprovisional[implicitcount])
                    contadorcolumna = contadorcolumna + 1
                    implicitcount = implicitcount + 1
                contadorfila = contadorfila + 1
                ws2 = workbook.create_sheet("Another Name", c + 1)
                ws2.title = arregloipinstancia[c]
                ws2.column_dimensions["A"].width = 15
                ws2.column_dimensions["B"].width = 10
                ws2.column_dimensions["C"].width = 15
                ws2.column_dimensions["D"].width = 25
                ws2.column_dimensions["E"].width = 20
                ws2.column_dimensions["F"].width = 25
                ws2[f'{"A"}{1}'] = "IP PE"##########resumen ejecutivo
                ws2[f'{"B"}{1}'] = "Instancia"
                ws2[f'{"C"}{1}'] = "Contador"        ##########pestañas
                ws2[f'{"D"}{1}'] = "Ultimo TCN"
                ws2[f'{"E"}{1}'] = "Interfaz" 
                ws2[f'{"F"}{1}'] = "Ultima recolecion GMT"      
                ws2.auto_filter.ref = "A1:F1"    
                ws2["A1"].font = bold_font
                ws2["B1"].font = bold_font
                ws2["C1"].font = bold_font
                ws2["D1"].font = bold_font
                ws2["E1"].font = bold_font
                ws2["F1"].font = bold_font
                ws2["A1"].font = white_font
                ws2["B1"].font = white_font
                ws2["C1"].font = white_font
                ws2["D1"].font = white_font
                ws2["E1"].font = white_font
                ws2["F1"].font = white_font
                ws2["A1"].fill = clarofill
                ws2["B1"].fill = clarofill
                ws2["C1"].fill = clarofill
                ws2["D1"].fill = clarofill
                ws2["E1"].fill = clarofill
                ws2["F1"].fill = clarofill
                for row_index, row_data in enumerate(arreglo[c], 1): 
                #print (arreglo[c])
                    col_letter = get_column_letter(contadorcolumna)
               # ws2 = wb.create_sheet("Another Name", 0)
                    #title = "My sheet name"
                    #sheet.title = arregloipinstancia[c]
                    for col_index, cell_data in enumerate(row_data, 1):
                        col_letter = get_column_letter(col_index)
                        ws2[f'{col_letter}{row_index + 1}'] = cell_data
        c = c + 1
    nombrearchivo = nombrearchivo + ".xlsx"
    workbook.save(nombrearchivo)
    return (arreglo)
tcnreport()

